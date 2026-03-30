"""
Apply NFC (Note For Claude) revisions to WIP_NFL_Survey_v0.xlsx
Based on RT3 review comments captured 2026-03-30.

Columns:
  3  = Question ID
  5  = Question text
  7  = Guide Explanation
  10 = Answer Options Definitions (pipe-separated)
  13 = Review Comment (added by export)
  14 = Review Flag
"""
import openpyxl
from pathlib import Path

XLSX = Path(__file__).parent.parent / "working/survey/WIP_NFL_Survey_v0.xlsx"
SHEET = "Combined Survey Questions"

# ─── Revision data ────────────────────────────────────────────────────────────
# Keys = Question ID as it appears in the sheet
# Values = dict with any of: question, options, guide
# "options" is the pipe-separated Answer Options Definitions string

REVISIONS = {

    # ── TI layer ──────────────────────────────────────────────────────────────

    "TI_1": {
        "question": (
            "How many Wi-Fi access points (APs) are currently deployed "
            "throughout your venue, per 1,000 seats?"
        ),
        "options": (
            "A: Fewer than 3 APs per 1,000 seats — significantly under-deployed, "
            "high likelihood of fan connectivity issues at capacity | "
            "B: 3–5 APs per 1,000 seats — below NFL median; fan experience will "
            "degrade during peak game-day use | "
            "C: 6–8 APs per 1,000 seats — industry standard for established NFL venues | "
            "D: 9–12 APs per 1,000 seats — above average; consistent performance "
            "under high-density fan load | "
            "E: More than 12 APs per 1,000 seats — industry-leading density; "
            "supports next-generation applications and future growth | "
            "Data not available / unknown | "
            "We don't use this technology today"
        ),
        "guide": (
            "AP density is the most fundamental indicator of wireless capacity. "
            "Leading NFL venues average 8–12 APs per 1,000 seats. "
            "To calculate: divide total deployed APs by venue seating capacity, "
            "then multiply by 1,000. Example: 700 APs at a 70,000-seat venue = "
            "10 APs per 1,000 seats. Under-densified Wi-Fi is the primary cause "
            "of poor in-venue connectivity regardless of technology generation."
        ),
    },

    "TI_3": {
        "question": (
            "What is the average measured Wi-Fi download speed available "
            "to fans during peak game-day use?"
        ),
        "options": (
            "A: Under 10 Mbps measured during peak game-day — "
            "below minimum threshold for modern device use | "
            "B: 10–24 Mbps measured during peak game-day — "
            "below NFL median; acceptable for basic use but limiting for video | "
            "C: 25–50 Mbps measured during peak game-day — "
            "industry median for established NFL venues | "
            "D: 51–150 Mbps measured during peak game-day — "
            "above average; supports streaming and high-demand applications | "
            "E: Over 150 Mbps measured during peak game-day — "
            "industry-leading performance | "
            "We have not measured fan-facing Wi-Fi speed during game day | "
            "Data not available / unknown"
        ),
        "guide": (
            "Measured throughput under game-day load is the ground truth of "
            "network quality — not vendor-specified or lab speeds. "
            "Peak game-day measurement (highest concurrent-user event) is the "
            "meaningful metric. The NFL venue median is 25–50 Mbps. "
            "Inability to measure is itself a low-maturity signal and should "
            "be recorded as a non-response (score 0). "
            "Cross-reference TI_1 (AP density) and TI_6 (uplink) — "
            "low speed despite high AP density suggests backhaul or "
            "configuration issues."
        ),
    },

    "TI_6": {
        "question": (
            "What is your primary internet uplink capacity into the venue on game days?"
        ),
        "options": (
            "A: Under 5 Gbps — significantly below capacity for a full stadium event | "
            "B: 5–9 Gbps — below the NFL median; adequate for moderate loads "
            "but likely to constrain peak demand | "
            "C: 10–20 Gbps — typical for an established NFL venue under "
            "current operational workloads | "
            "D: 21–40 Gbps — above average; supports high-bandwidth "
            "applications and provides headroom for growth | "
            "E: More than 40 Gbps — industry-leading capacity, "
            "typical of the most modern NFL venues | "
            "Data not available / unknown"
        ),
        "guide": (
            "Total internet uplink capacity determines the ceiling on all "
            "cloud-dependent services during an event — fan Wi-Fi, broadcast "
            "feeds, POS, access control, and mobile app traffic all compete "
            "for this pipe. The NFL venue median is 10–20 Gbps. "
            "Venues under 5 Gbps typically experience significant degradation "
            "at capacity. Note: 'unknown' is a non-response (score 0)."
        ),
    },

    "TI_7": {
        "question": (
            "How resilient is your venue's internet connectivity if the "
            "primary circuit fails during a live event?"
        ),
        "options": (
            "A: No redundancy — a single internet connection is in use "
            "with no backup or failover path | "
            "B: A secondary circuit exists but failover requires manual "
            "intervention by IT staff and takes several minutes to activate | "
            "C: Dual ISP paths are in place with a documented manual "
            "failover procedure that has been tested at least once | "
            "D: Dual ISP paths are configured for automatic failover "
            "with tested switching under game-day load conditions | "
            "E: Fully automated failover with active load balancing, "
            "real-time circuit health verification, documented SLAs with ISP "
            "partners, and regular disaster recovery testing | "
            "Data not available / unknown"
        ),
        "guide": (
            "Internet redundancy is a critical-path infrastructure requirement "
            "for any venue running cloud-dependent applications (ticketing, POS, "
            "mobile ordering, Wi-Fi authentication). A single circuit failure "
            "during peak game-day load can disable multiple fan-facing systems "
            "simultaneously. The minimum acceptable standard for a modern NFL "
            "venue is dual ISP paths with automatic failover (option D). "
            "Venues at option A or B face unacceptable operational risk."
        ),
    },

    "TI_8": {
        "question": (
            "How is your in-venue network segmented across user groups and "
            "use cases, and how is bandwidth allocated between them?"
        ),
        "options": (
            "A: No segmentation — all users and systems share a single flat network | "
            "B: Basic separation — fan traffic is isolated from staff and "
            "back-of-house operations on separate networks | "
            "C: Multiple dedicated segments — fan, staff, media, and venue "
            "operations each run on separate logical networks (VLANs) | "
            "D: Segmented networks with defined bandwidth allocation — "
            "Quality of Service (QoS) policies prioritize critical systems "
            "such as POS, gate access, and broadcast over general fan traffic | "
            "E: Fully segmented architecture with dynamic traffic shaping — "
            "bandwidth is allocated in real time based on demand, critical systems "
            "are guaranteed capacity, and network performance is actively "
            "monitored and reported during events | "
            "Data not available / unknown"
        ),
        "guide": (
            "Network segmentation determines both service reliability and fan "
            "experience consistency. Flat (unsegmented) networks allow "
            "non-critical traffic to compete with and degrade POS, access "
            "control, and broadcast systems during peak load. "
            "Note: network security posture (micro-segmentation, zero-trust "
            "access controls, threat detection) is a separate dimension not "
            "measured here — see the dedicated network security question."
        ),
    },

    "TI_11": {
        "question": (
            "How does your organization manage and prioritize Wi-Fi traffic "
            "to deliver a consistent fan experience during peak game-day demand?"
        ),
        "options": (
            "A: No traffic management in place — all devices compete equally "
            "for available bandwidth with no prioritization | "
            "B: Basic traffic shaping applied — fan traffic is separated from "
            "operational systems but no prioritization rules exist within fan traffic | "
            "C: Quality of Service (QoS) policies are configured to prioritize "
            "critical applications such as mobile ticketing, POS, and mobile "
            "ordering above general browsing and streaming | "
            "D: Dynamic QoS policies are actively managed during events — "
            "traffic is monitored in real time and policies adjusted based on "
            "congestion patterns, with defined thresholds for fan-facing "
            "application performance | "
            "E: Fully automated, policy-driven traffic management — "
            "fan application performance targets are defined and enforced, "
            "congestion is predicted and mitigated before fan impact occurs, "
            "and post-event reports validate adherence to performance baselines | "
            "Data not available / unknown"
        ),
        "guide": (
            "Traffic prioritization is what turns raw AP density and uplink "
            "capacity into a consistent fan experience. A venue may have "
            "adequate hardware but still deliver poor connectivity if all "
            "traffic competes equally. This question measures whether the "
            "organization actively manages the quality of experience for fans "
            "vs. simply providing connectivity. Cross-reference TI_1 "
            "(AP density), TI_6 (uplink), and TI_8 (segmentation)."
        ),
    },

    "TI_14": {
        "question": (
            "How consistent and standardized is your venue's network "
            "infrastructure hardware, and how was it procured and deployed?"
        ),
        "options": (
            "A: Infrastructure has been built piecemeal across multiple "
            "generations — hardware from different vendors and eras is in "
            "use with no overarching standards; significant compatibility "
            "gaps or unsupported equipment exist | "
            "B: Some standardization exists within individual network tiers "
            "or areas of the venue, but the overall environment remains mixed "
            "with legacy hardware still in use across meaningful portions of "
            "the infrastructure | "
            "C: A venue-wide hardware standard is in place for primary "
            "infrastructure — the same vendor and platform covers most of "
            "the network, though some legacy equipment from earlier builds "
            "remains in secondary areas | "
            "D: Consistent hardware standards are enforced across all network "
            "tiers — procurement follows a formal evaluation and approval "
            "process, refresh cycles are documented, and configurations are "
            "centrally managed | "
            "E: Infrastructure is fully standardized, purpose-engineered for "
            "venue scale, and operated as a unified platform — hardware meets "
            "defined specifications across all layers, configurations are "
            "version-controlled, and the environment is designed to accommodate "
            "future technology additions without requiring major re-architecture | "
            "Data not available / unknown"
        ),
    },

    # ── DI layer ──────────────────────────────────────────────────────────────

    "DI_4": {
        "question": (
            "How does your organization leverage real-time data streaming "
            "capabilities to enhance the fan experience and support "
            "game-day operations?"
        ),
        "options": (
            "A: No real-time data streaming capabilities in place — "
            "data is captured and reviewed after events in batch reports | "
            "B: Limited real-time data is available for one or two isolated "
            "use cases (such as gate entry counts or POS totals), but it is "
            "not integrated into operational decisions during the event | "
            "C: Real-time data streams from multiple systems are available "
            "to operations teams during the event and are used to make "
            "reactive adjustments such as re-routing fans or opening "
            "additional POS stations | "
            "D: Real-time data is integrated across fan-facing and operational "
            "systems — streaming data informs both fan experience decisions "
            "(dynamic signage content, in-app notifications) and operational "
            "responses (staffing adjustments, queue management) during the event | "
            "E: A unified real-time streaming platform ingests data from all "
            "major venue systems, feeds automated operational responses and "
            "personalized fan interactions during the event, and provides "
            "operations leadership with a live performance dashboard "
            "throughout the game | "
            "Data not available / unknown | "
            "We don't use this technology today"
        ),
    },

    "DI_5": {
        "question": (
            "How mature and scalable is your organization's API infrastructure "
            "and system integration architecture?"
        ),
        "options": (
            "A: APIs exist for some systems but were built ad hoc — "
            "integrations are point-to-point, poorly documented, and require "
            "manual intervention when systems change | "
            "B: Several integrations are functional, but there is no centralized "
            "management layer — each integration is managed independently, "
            "creating inconsistent reliability and limited visibility | "
            "C: A structured integration layer is in place — an API gateway or "
            "middleware platform manages connections between core systems, with "
            "documented interfaces, consistent authentication, and basic monitoring | "
            "D: API infrastructure is mature — a governed API gateway handles "
            "all major system integrations, rate limits and SLAs are enforced, "
            "developer documentation is maintained, and integration health is "
            "monitored in near real time | "
            "E: API-first architecture is fully adopted — all major systems "
            "expose versioned APIs, a platform team owns integration governance, "
            "new systems are evaluated for API compatibility before procurement, "
            "and the integration layer enables rapid onboarding of new "
            "capabilities without disrupting existing connections | "
            "Data not available / unknown"
        ),
    },

    "DI_6": {
        "question": (
            "How would you describe the overall architecture of your "
            "organization's core venue technology applications and platforms?"
        ),
        "options": (
            "A: Primarily monolithic or legacy applications — core systems "
            "are large, tightly coupled platforms that are difficult to update, "
            "scale, or integrate without significant effort | "
            "B: Some modernization has occurred — a mix of legacy and newer "
            "standalone applications, but systems are largely independent "
            "and integrations are limited | "
            "C: Applications are increasingly modular — key systems have "
            "been decomposed into independent services or components, enabling "
            "more flexible updates and integrations, though not fully standardized | "
            "D: Modern application architecture is broadly adopted — most "
            "systems are built or configured as loosely coupled services, "
            "containerized where appropriate, and deployed through automated "
            "pipelines with defined integration standards | "
            "E: Cloud-native, event-driven architecture is the organizational "
            "standard — services are independently deployable, infrastructure "
            "is managed as code, and event-driven patterns enable real-time "
            "system communication across the venue technology stack | "
            "Data not available / unknown"
        ),
    },

    "DI_7": {
        "options": (
            "Multi-factor authentication (MFA) required for all staff accessing "
            "venue technology systems | "
            "Single Sign-On (SSO) in place so staff use one set of credentials "
            "across core venue applications | "
            "Role-based access control (RBAC) enforced — staff access only "
            "the systems and data their role requires | "
            "Privileged access management (PAM) in place for administrator "
            "and system-level accounts | "
            "Automated provisioning and de-provisioning — access is granted "
            "and revoked automatically when staff are onboarded or leave | "
            "Access certification in place — access rights are reviewed and "
            "confirmed on a regular schedule (e.g., quarterly) | "
            "Zero Trust access model applied — all access requests are "
            "verified regardless of network location | "
            "Machine-to-machine identity management for system integrations "
            "and APIs | "
            "Centralized identity logging and auditing — all access events "
            "are logged and available for review | "
            "Vendor and third-party access is controlled through time-limited, "
            "monitored sessions | "
            "None of the above"
        ),
    },

    "DI_10": {
        "question": (
            "How many self-service kiosks (food & beverage ordering, ticketing, "
            "or information) are currently operational in your venue, "
            "per 10,000 seats?"
        ),
        "options": (
            "A: No self-service kiosks are currently operational | "
            "B: Fewer than 5 kiosks per 10,000 seats — minimal deployment, "
            "limited self-service coverage | "
            "C: 5–10 kiosks per 10,000 seats — moderate deployment; "
            "self-service available in key areas | "
            "D: 11–20 kiosks per 10,000 seats — broad coverage across "
            "concourse and concession zones | "
            "E: More than 20 kiosks per 10,000 seats — high-density deployment; "
            "self-service accessible across all fan-facing areas | "
            "Data not available / unknown | "
            "We don't use this technology today"
        ),
        "guide": (
            "Normalize kiosk count by venue capacity for cross-club comparability. "
            "Example: 42 kiosks at a 70,000-seat venue = 6 kiosks per 10,000 seats (option C). "
            "Note: clubs where concessions are fully managed by a third-party concessionaire "
            "may not have direct visibility into kiosk counts — flag this in "
            "the concessionaire response option and confirm with the respondent."
        ),
    },

    "DI_11": {
        "question": (
            "Approximately how many surveillance and operational cameras "
            "are currently deployed throughout your venue "
            "(interior, concourse, and exterior combined)?"
        ),
        "options": (
            "A: Fewer than 100 cameras — minimal coverage, limited to "
            "primary entry and high-risk locations only | "
            "B: 100–250 cameras — partial coverage; key areas monitored "
            "but significant gaps remain | "
            "C: 251–450 cameras — standard coverage for an NFL venue; "
            "all primary fan-facing areas monitored | "
            "D: 451–700 cameras — above average; comprehensive coverage "
            "including all concourse, concession, and exterior zones | "
            "E: More than 700 cameras — high-density deployment; supports "
            "advanced analytics and AI-driven use cases | "
            "Data not available / unknown"
        ),
    },

    "DI_12": {
        "question": (
            "Which of the following edge and content delivery capabilities "
            "does your organization currently have deployed? "
            "(Select all that apply)"
        ),
        "options": (
            "Content Delivery Network (CDN) for fan-facing digital content "
            "(website, mobile app, or streaming media) | "
            "Edge caching for mobile app assets to reduce load times "
            "inside the venue | "
            "In-venue edge compute nodes deployed for low-latency "
            "application processing | "
            "Distributed ticketing validation at gates — gate scanning "
            "processes tickets locally without requiring a cloud round-trip | "
            "Edge-based POS transaction processing that remains functional "
            "if internet connectivity is interrupted | "
            "In-venue media distribution system for local delivery of "
            "video and audio content | "
            "Geo-distributed redundancy so fan-facing services remain "
            "available if a primary data center goes offline | "
            "None of the above"
        ),
    },

    # ── TS layer ──────────────────────────────────────────────────────────────

    "TS_2": {
        "question": (
            "How deeply integrated and operationally mature is your "
            "organization's ticketing and access control implementation?"
        ),
        "options": (
            "A: Basic ticketing platform in place for sales only — access "
            "control is managed separately and the two systems do not share data | "
            "B: Ticketing and access control are connected — digital ticket "
            "scanning is in use at gates, but entry data is not integrated "
            "into other venue systems | "
            "C: Ticketing data is integrated with at least one other operational "
            "system (such as CRM or game-day operations) — entry data is "
            "available for post-event reporting | "
            "D: Ticketing is part of a connected venue platform — entry data "
            "flows in real time to operations, fan identity is linked to ticket "
            "purchase, and the platform supports mobile and contactless entry | "
            "E: Fully integrated ticketing ecosystem — real-time entry data "
            "informs staffing decisions, fan identity connects purchase to "
            "in-venue behavior, and the platform enables data-driven fan "
            "engagement before, during, and after the event | "
            "Data not available / unknown | "
            "We don't use this technology today"
        ),
        "guide": (
            "This question measures the depth of ticketing implementation, "
            "not the number of platforms in use. Having multiple ticketing "
            "vendors typically signals fragmentation rather than maturity. "
            "Score is based on integration depth and operational use of "
            "ticketing data, not vendor selection. "
            "Note: clubs where ticketing is fully managed by the league or "
            "a concessionaire may have limited direct visibility — "
            "confirm scope with the respondent."
        ),
    },

    "TS_3": {
        "question": (
            "How broadly are point-of-sale and payment capabilities "
            "deployed and accessible to fans throughout your venue?"
        ),
        "options": (
            "A: POS is available only at fixed traditional concession stands — "
            "fans must travel to a counter and wait in line to complete "
            "a purchase anywhere in the venue | "
            "B: POS coverage extends to the majority of concession locations, "
            "but all transactions require visiting a fixed station | "
            "C: Mobile POS or line-busting devices are deployed in "
            "high-traffic areas, allowing staff to process transactions "
            "outside of fixed counters and reduce wait times | "
            "D: Fans can order and pay via mobile app for pickup at "
            "designated express windows in most areas of the venue — "
            "reducing the need to leave their seat for most transactions | "
            "E: Full seat-level commerce is available — fans can order "
            "food, beverages, and merchandise directly from their seat "
            "via mobile app or in-seat device, with delivery throughout "
            "the venue | "
            "Data not available / unknown | "
            "We don't use this technology today"
        ),
        "guide": (
            "This question measures POS accessibility and fan convenience, "
            "not the number or brand of POS systems deployed. Having more "
            "POS vendors does not increase the score. "
            "Note: clubs where concessions are managed by a third-party "
            "concessionaire may have limited visibility into POS coverage — "
            "confirm scope with the respondent and note concessionaire "
            "arrangement where applicable."
        ),
    },

    "TS_9": {
        "options": (
            "A: Under 25% of transactions are cashless — cash remains "
            "the dominant payment method across the venue | "
            "B: 25–49% of transactions are cashless — cashless options "
            "are available but cash is still widely used | "
            "C: 50–69% of transactions are cashless — cashless is the "
            "majority payment method, consistent with the current NFL median | "
            "D: 70–89% of transactions are cashless — the venue operates "
            "primarily cashless with minimal cash handling infrastructure remaining | "
            "E: 90% or more of transactions are cashless — the venue "
            "operates as effectively cashless, consistent with fully cashless "
            "NFL venues | "
            "Data not available / unknown"
        ),
    },

    # ── AI layer ──────────────────────────────────────────────────────────────

    "AI_3": {
        "question": (
            "Which of the following computer vision applications does your "
            "organization currently use during game-day operations? "
            "(Select all that apply)"
        ),
        "options": (
            "Crowd density monitoring — cameras analyze attendance distribution "
            "in real time to identify overcrowding or underutilized areas | "
            "Queue and wait time estimation — computer vision measures line "
            "lengths at concessions, gates, or restrooms | "
            "Entry and gate management — automated camera systems identify "
            "or count fans moving through entry points | "
            "Security threat detection — computer vision flags anomalous "
            "behavior or restricted zone breaches for security staff | "
            "Cashierless checkout — cameras and sensors enable "
            "frictionless checkout-free transactions | "
            "Fan sentiment or engagement detection — computer vision "
            "analyzes crowd response during key game moments | "
            "Sponsorship impression measurement — cameras measure fan "
            "exposure to sponsor placements or displays | "
            "Operational monitoring — cameras detect equipment issues, "
            "spills, or maintenance needs in real time | "
            "None of the above — we do not currently use computer vision "
            "during game-day operations"
        ),
    },

    "AI_5": {
        "question": (
            "Which of the following Generative AI applications is your "
            "organization currently running in production? "
            "(Select all that apply)"
        ),
        "options": (
            "Fan-facing chatbot or virtual assistant powered by a large "
            "language model (LLM) — handles inquiries during or around events | "
            "Dynamic in-venue or social media content generated by AI in "
            "real time (highlight summaries, signage copy, game-day posts) | "
            "AI-generated personalized offers or promotions delivered to "
            "fans during the event via app or messaging | "
            "Natural language querying of operational data — staff ask "
            "questions of venue systems in plain language | "
            "AI-generated post-game content (recaps, sponsor reports, "
            "fan communications) | "
            "AI-assisted workforce management or real-time staff deployment "
            "recommendations during events | "
            "None of the above — we are evaluating or piloting Generative AI "
            "but nothing is in production"
        ),
    },

    "AI_7": {
        "question": (
            "How clearly defined and resourced is your organization's "
            "roadmap for AI investment and deployment?"
        ),
        "options": (
            "A: No formal AI roadmap — AI initiatives are evaluated "
            "opportunistically with no defined investment commitment or timeline | "
            "B: An AI roadmap or vision has been drafted but lacks committed "
            "capital allocation — priorities exist on paper but no budget "
            "has been formally approved | "
            "C: AI investment has been approved at a program level — "
            "a defined budget and prioritized use case list exists, though "
            "delivery timelines and resourcing are not fully locked | "
            "D: A funded AI roadmap is in place with committed capital, "
            "assigned owners, and defined delivery milestones — progress "
            "is reviewed regularly and use cases are sequenced by business impact | "
            "E: AI investment is a strategic capability commitment — "
            "multi-year capital allocation is approved, a dedicated team "
            "owns execution, the roadmap is tied to measurable business "
            "outcomes, and governance structures ensure accountability "
            "for delivery against plan | "
            "Data not available / unknown"
        ),
    },
}

# Questions to mark as DROP (flagged by reviewer — not deleted, marked for team review)
DROP_IDS = {"DI_13", "DI_14", "DI_15", "TS_10", "AI_10"}
DROP_NOTE = "MARKED FOR DROP — reviewer confirmed not needed in final survey. Remove after Vik review."


def apply():
    wb = openpyxl.load_workbook(str(XLSX))
    ws = wb[SHEET]

    updated = []
    dropped = []

    # Build QID → row index
    qid_to_row = {}
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=3).value
        if val and str(val).strip() not in ("", "Question ID"):
            qid_to_row[str(val).strip()] = r

    # Apply revisions
    for qid, changes in REVISIONS.items():
        row = qid_to_row.get(qid)
        if not row:
            print(f"  [WARN] {qid} not found in sheet — skipping")
            continue
        if "question" in changes:
            ws.cell(row=row, column=5, value=changes["question"])
        if "options" in changes:
            ws.cell(row=row, column=10, value=changes["options"])
        if "guide" in changes:
            ws.cell(row=row, column=7, value=changes["guide"])
        updated.append(qid)
        print(f"  [OK] {qid} updated")

    # Mark drops
    for qid in DROP_IDS:
        row = qid_to_row.get(qid)
        if row:
            existing = ws.cell(row=row, column=13).value or ""
            if DROP_NOTE not in str(existing):
                ws.cell(row=row, column=13, value=DROP_NOTE)
            dropped.append(qid)
            print(f"  [DROP] {qid} marked")

    wb.save(str(XLSX))
    print(f"\nDone. {len(updated)} questions updated, {len(dropped)} marked for drop.")
    print(f"Saved: {XLSX}")


if __name__ == "__main__":
    print(f"Applying NFC revisions to {XLSX.name} …\n")
    apply()
