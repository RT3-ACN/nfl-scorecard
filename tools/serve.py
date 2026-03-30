#!/usr/bin/env python3
"""
NFL Survey Tools — Standalone Server
Run from the NFL/ directory:   python3 tools/serve.py
Or from NFL/tools/:             python3 serve.py

Opens automatically at http://localhost:9876
Requires: openpyxl  (pip install openpyxl)
"""
import datetime
import http.server
import json
import os
import threading
import webbrowser
from pathlib import Path

PORT           = 9876
TOOLS_DIR      = Path(__file__).parent
NFL_DIR        = TOOLS_DIR.parent
DELIVERABLES_DIR = NFL_DIR / "deliverables"
SURVEY_DIR     = NFL_DIR / "working" / "survey"
SURVEY_XLSX    = SURVEY_DIR / "WIP_NFL_Survey_v0.xlsx"
COMMENTS_FILE  = SURVEY_DIR / "nfl_review_comments.json"
KANBAN_FILE    = TOOLS_DIR / "kanban-state.json"
DASHBOARD_HTML = DELIVERABLES_DIR / "nfl-scorecard-dashboard.html"
MEETINGS_DIR   = NFL_DIR / "meetings"

SURVEY_SHEET  = "Combined Survey Questions"

KNOWN_ISSUES = {
    "DI_3": "Count normalization needed — absolute counts not comparable across venue sizes",
    "DI_5": "'We have not measured' must be a non-response (score 0), not option A",
    "TS_3": "Must address concessionaire-managed POS scenario",
}

MIME = {
    ".html": "text/html; charset=utf-8",
    ".js":   "application/javascript",
    ".css":  "text/css",
    ".json": "application/json",
    ".ico":  "image/x-icon",
}

_q_cache = None
_q_mtime = 0.0


# ── DATA FUNCTIONS ─────────────────────────────────────────────────────────────

def parse_questions() -> list:
    global _q_cache, _q_mtime
    if not SURVEY_XLSX.exists():
        return []
    try:
        mtime = SURVEY_XLSX.stat().st_mtime
        if _q_cache is not None and mtime == _q_mtime:
            return _q_cache
        import openpyxl
        wb = openpyxl.load_workbook(str(SURVEY_XLSX), data_only=True)
        ws = wb[SURVEY_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(c).strip() if c else "" for c in rows[0]]

        def col(row, name, fallback=""):
            try:
                v = row[headers.index(name)]
                return str(v).strip() if v else fallback
            except (ValueError, IndexError):
                return fallback

        questions = []
        for row in rows[1:]:
            qid = col(row, "Question ID")
            if not qid:
                continue
            opts_raw = col(row, "Answer Options Definitions")
            opts = [o.strip() for o in opts_raw.split("|") if o.strip()] if opts_raw else []
            questions.append({
                "id":           qid,
                "category":     col(row, "Survey Category"),
                "domain":       col(row, "Domain Covered"),
                "question":     col(row, "Question"),
                "guide":        col(row, "Guide Explanation"),
                "format":       col(row, "Answer Format"),
                "style":        col(row, "Answer Options Style"),
                "options":      opts,
                "answer_guide": col(row, "Answer Guide"),
                "revision_note": col(row, "Revision Note on Scorability"),
                "known_issue":  KNOWN_ISSUES.get(qid, ""),
            })
        wb.close()
        _q_cache = questions
        _q_mtime = mtime
        return questions
    except Exception as e:
        print(f"[ERROR] parse_questions: {e}")
        return []


def parse_revision_notes() -> list:
    """Return questions that have a non-empty Revision Note on Scorability."""
    return [q for q in parse_questions() if q.get("revision_note", "").strip()]


def load_comments() -> dict:
    if not COMMENTS_FILE.exists():
        return {"meta": {"source": SURVEY_XLSX.name,
                         "created": datetime.date.today().isoformat()},
                "comments": {}}
    try:
        return json.loads(COMMENTS_FILE.read_text())
    except Exception:
        return {"comments": {}}


def save_comment(qid: str, flag: str, comment: str) -> dict:
    data = load_comments()
    data.setdefault("comments", {})[qid] = {
        "flag":      flag,
        "comment":   comment,
        "timestamp": datetime.datetime.now().isoformat()[:19],
    }
    data.setdefault("meta", {})["last_updated"] = datetime.date.today().isoformat()
    COMMENTS_FILE.write_text(json.dumps(data, indent=2))
    return {"ok": True}


def export_to_excel() -> dict:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "msg": "openpyxl not installed — run: pip install openpyxl"}
    if not SURVEY_XLSX.exists():
        return {"ok": False, "msg": f"Survey file not found: {SURVEY_XLSX}"}
    try:
        comments = load_comments().get("comments", {})
        wb = openpyxl.load_workbook(str(SURVEY_XLSX))
        ws = wb["NFL Claude V0"]
        ws.cell(row=7, column=10, value="Review Comment")
        ws.cell(row=7, column=11, value="Review Flag")
        for row in range(8, ws.max_row + 1):
            qid = ws.cell(row=row, column=2).value
            if qid and str(qid).strip() in comments:
                c = comments[str(qid).strip()]
                ws.cell(row=row, column=10, value=c.get("comment", ""))
                ws.cell(row=row, column=11, value=c.get("flag", ""))
        wb.save(str(SURVEY_XLSX))
        return {"ok": True, "file": SURVEY_XLSX.name}
    except Exception as e:
        return {"ok": False, "msg": str(e)}


def _parse_frontmatter(content: str):
    """Return (meta_dict, body_str). Handles --- YAML frontmatter."""
    if not content.startswith("---\n"):
        return {}, content
    end = content.find("\n---\n", 4)
    if end == -1:
        return {}, content
    meta = {}
    for line in content[4:end].split("\n"):
        if ":" in line:
            k, _, v = line.partition(":")
            meta[k.strip()] = v.strip().strip("'\"")
    return meta, content[end + 5:]


def list_meetings() -> list:
    if not MEETINGS_DIR.exists():
        return []
    result = []
    for f in sorted(MEETINGS_DIR.glob("*.md"), reverse=True):
        content = f.read_text(encoding="utf-8")
        meta, body = _parse_frontmatter(content)
        title = next((l.lstrip("# ").strip() for l in body.split("\n") if l.startswith("# ")), f.stem)
        date = meta.get("date", f.stem[:10] if len(f.stem) >= 10 and f.stem[4] == "-" else "")
        result.append({
            "slug":    f.stem,
            "title":   title,
            "date":    date,
            "type":    meta.get("type", "internal"),
            "source":  meta.get("source", "manual"),
            "content": body,
        })
    return result


def save_meeting(title: str, date: str, mtype: str, content: str) -> dict:
    import re, time
    slug = re.sub(r"[^a-z0-9]+", "-", f"{date}-{title}".lower()).strip("-")
    path = MEETINGS_DIR / f"{slug}.md"
    if path.exists():
        slug += f"-{int(time.time())}"
        path = MEETINGS_DIR / f"{slug}.md"
    MEETINGS_DIR.mkdir(exist_ok=True)
    fm = f"---\ndate: {date}\ntype: {mtype}\nsource: manual\n---\n\n"
    path.write_text(fm + f"# {title}\n**Date:** {date}\n**Type:** {mtype.capitalize()}\n\n{content}", encoding="utf-8")
    return {"ok": True, "slug": slug}


def load_kanban() -> dict:
    if not KANBAN_FILE.exists():
        return _default_kanban()
    try:
        return json.loads(KANBAN_FILE.read_text())
    except Exception:
        return _default_kanban()


def save_kanban(data: dict) -> dict:
    KANBAN_FILE.write_text(json.dumps(data, indent=2))
    return {"ok": True}


def _default_kanban() -> dict:
    return {
        "columns": [
            {"id": "backlog",    "title": "Backlog",       "cards": []},
            {"id": "inprogress", "title": "In Progress",   "cards": []},
            {"id": "review",     "title": "In Review",     "cards": []},
            {"id": "done",       "title": "Done",          "cards": []},
        ]
    }


# ── HTTP HANDLER ───────────────────────────────────────────────────────────────

class Handler(http.server.BaseHTTPRequestHandler):

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def do_GET(self):
        p = self.path.split("?")[0]

        if p in ("/", "/index.html", ""):
            self._file(DELIVERABLES_DIR / "nfl-scorecard-dashboard.html")
        elif p == "/review":
            self._file(TOOLS_DIR / "review.html")
        elif p == "/kanban":
            self._file(TOOLS_DIR / "kanban.html")
        elif p == "/changes":
            self._file(TOOLS_DIR / "changes.html")
        elif p == "/api/questions":
            self._json(parse_questions())
        elif p == "/api/revision-notes":
            self._json(parse_revision_notes())
        elif p == "/api/comments":
            self._json(load_comments())
        elif p == "/api/meetings":
            self._json(list_meetings())
        elif p == "/api/kanban":
            self._json(load_kanban())
        else:
            # Try deliverables/ first, then tools/ for assets
            for base in (DELIVERABLES_DIR, TOOLS_DIR):
                target = base / p.lstrip("/")
                if target.exists() and target.is_file():
                    self._file(target)
                    return
            self.send_response(404)
            self.end_headers()

    def do_POST(self):
        length = int(self.headers.get("Content-Length", 0))
        body   = self.rfile.read(length)
        p      = self.path.split("?")[0]

        if p == "/api/comments":
            try:
                d = json.loads(body)
                self._json(save_comment(d.get("id",""), d.get("flag",""), d.get("comment","")))
            except Exception as e:
                self._json({"error": str(e)}, 500)
        elif p == "/api/export":
            self._json(export_to_excel())
        elif p == "/api/kanban":
            try:
                self._json(save_kanban(json.loads(body)))
            except Exception as e:
                self._json({"error": str(e)}, 500)
        elif p == "/api/meetings":
            try:
                d = json.loads(body)
                self._json(save_meeting(d.get("title","Untitled"), d.get("date",""), d.get("type","internal"), d.get("content","")))
            except Exception as e:
                self._json({"error": str(e)}, 500)
        else:
            self._json({"error": "Not found"}, 404)

    def _file(self, path: Path):
        if not path.exists():
            self.send_response(404); self.end_headers(); return
        data = path.read_bytes()
        ct   = MIME.get(path.suffix, "application/octet-stream")
        self.send_response(200)
        self._cors()
        self.send_header("Content-Type", ct)
        self.send_header("Content-Length", str(len(data)))
        self.send_header("Cache-Control", "no-cache")
        self.end_headers()
        self.wfile.write(data)

    def _json(self, data, status=200):
        payload = json.dumps(data).encode()
        self.send_response(status)
        self._cors()
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(payload)))
        self.end_headers()
        self.wfile.write(payload)

    def _cors(self):
        self.send_header("Access-Control-Allow-Origin",  "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")

    def log_message(self, *_):
        pass


# ── MAIN ───────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    os.chdir(TOOLS_DIR)
    url = f"http://localhost:{PORT}"
    print(f"\n  NFL Survey Tools")
    print(f"  ─────────────────────────────")
    print(f"  Home             →  {url}")
    print(f"  Question Review  →  {url}/review")
    print(f"  Revision Notes   →  {url}/changes")
    print(f"  Kanban Board     →  {url}/kanban")
    print(f"  Ctrl+C to stop\n")

    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print(f"  ⚠  openpyxl not installed — Question Review will show no questions.")
        print(f"     Fix: pip install openpyxl\n")

    if not SURVEY_XLSX.exists():
        print(f"  ⚠  Survey file not found: {SURVEY_XLSX}")
        print(f"     Question Review will not work until the file is present.\n")

    threading.Timer(0.8, lambda: webbrowser.open(url)).start()
    server = http.server.HTTPServer(("localhost", PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n  Stopped.")
