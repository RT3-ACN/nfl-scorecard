#!/usr/bin/env python3
"""
xlsx_diff.py — Compare two versions of WIP_NFL_Survey_v0.xlsx question-by-question.

Usage:
    python3 xlsx_diff.py --old <path> --new <path> [--sheet "Combined Survey Questions"] [--json]

Output: human-readable diff to stdout (or JSON with --json).
"""
import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

SHEET = "Combined Survey Questions"
COMPARE_FIELDS = [
    "Question", "Domain Covered", "Guide Explanation",
    "Answer Format", "Answer Options Style", "Answer Options Definitions",
    "Answer Guide", "V1 Notes", "V1 Comments",
]
NEW_ONLY_FIELDS = ["Revision Note on Scorability"]


def normalize_id(qid: str) -> str:
    """TI_01 → TI_1, TI_1 → TI_1"""
    m = re.match(r"([A-Za-z]+)_0*(\d+)", str(qid).strip())
    return f"{m.group(1).upper()}_{m.group(2)}" if m else str(qid).strip()


def load_questions(path: Path, sheet: str) -> tuple[list, dict]:
    wb = openpyxl.load_workbook(str(path), data_only=True)
    if sheet not in wb.sheetnames:
        sys.exit(f"Sheet '{sheet}' not found in {path}. Available: {wb.sheetnames}")
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(c).strip() if c else "" for c in rows[0]]
    questions = {}
    for row in rows[1:]:
        raw_id = row[headers.index("Question ID")] if "Question ID" in headers else row[2]
        if not raw_id or not str(raw_id).strip():
            continue
        norm = normalize_id(str(raw_id))
        questions[norm] = {
            "_orig_id": str(raw_id).strip(),
            **{h: (str(v).strip() if v else "") for h, v in zip(headers, row)},
        }
    wb.close()
    return headers, questions


def diff(old_path: Path, new_path: Path, sheet: str) -> dict:
    old_headers, old_qs = load_questions(old_path, sheet)
    new_headers, new_qs = load_questions(new_path, sheet)

    old_ids = set(old_qs)
    new_ids = set(new_qs)

    added_cols   = [c for c in new_headers if c not in old_headers and c]
    removed_cols = [c for c in old_headers if c not in new_headers and c]
    added_qs     = sorted(new_ids - old_ids)
    removed_qs   = sorted(old_ids - new_ids)

    modified = []
    for qid in sorted(old_ids & new_ids):
        o, n = old_qs[qid], new_qs[qid]
        field_diffs = {}
        for field in COMPARE_FIELDS:
            ov, nv = o.get(field, ""), n.get(field, "")
            if ov != nv:
                field_diffs[field] = {"old": ov, "new": nv}
        for field in NEW_ONLY_FIELDS:
            nv = n.get(field, "")
            if nv:
                field_diffs[field] = {"old": "", "new": nv}
        if field_diffs:
            modified.append({"id": qid, "orig_old": o["_orig_id"], "orig_new": n["_orig_id"], "fields": field_diffs})

    return {
        "structural": {
            "old_sheet_cols": len(old_headers),
            "new_sheet_cols": len(new_headers),
            "added_columns": added_cols,
            "removed_columns": removed_cols,
            "old_question_count": len(old_qs),
            "new_question_count": len(new_qs),
        },
        "added":    [{"id": qid, "question": new_qs[qid].get("Question", "")[:80]} for qid in added_qs],
        "removed":  [{"id": qid, "question": old_qs[qid].get("Question", "")[:80]} for qid in removed_qs],
        "modified": modified,
    }


def print_diff(result: dict):
    s = result["structural"]
    print("=" * 70)
    print("STRUCTURAL CHANGES")
    print("=" * 70)
    print(f"  Columns:   {s['old_sheet_cols']} → {s['new_sheet_cols']}")
    if s["added_columns"]:
        print(f"  Added cols:   {s['added_columns']}")
    if s["removed_columns"]:
        print(f"  Removed cols: {s['removed_columns']}")
    print(f"  Questions: {s['old_question_count']} → {s['new_question_count']}")

    if result["added"]:
        print(f"\n{'=' * 70}")
        print(f"ADDED QUESTIONS ({len(result['added'])})")
        print("=" * 70)
        for q in result["added"]:
            print(f"  + {q['id']}: {q['question']}")

    if result["removed"]:
        print(f"\n{'=' * 70}")
        print(f"REMOVED QUESTIONS ({len(result['removed'])})")
        print("=" * 70)
        for q in result["removed"]:
            print(f"  - {q['id']}: {q['question']}")

    if result["modified"]:
        print(f"\n{'=' * 70}")
        print(f"MODIFIED QUESTIONS ({len(result['modified'])})")
        print("=" * 70)
        for m in result["modified"]:
            id_note = f" (was {m['orig_old']})" if m["orig_old"] != m["id"] else ""
            print(f"\n  ~ {m['id']}{id_note}")
            for field, vals in m["fields"].items():
                print(f"    [{field}]")
                if vals["old"]:
                    print(f"      OLD: {vals['old'][:100]}")
                print(f"      NEW: {vals['new'][:100]}")

    print(f"\n{'=' * 70}")
    print(f"SUMMARY: +{len(result['added'])} added  -{len(result['removed'])} removed  ~{len(result['modified'])} modified")
    print("=" * 70)


def main():
    parser = argparse.ArgumentParser(description="Diff two NFL survey xlsx files.")
    parser.add_argument("--old",   required=True, help="Path to the old xlsx")
    parser.add_argument("--new",   required=True, help="Path to the new xlsx")
    parser.add_argument("--sheet", default=SHEET, help=f"Sheet name (default: '{SHEET}')")
    parser.add_argument("--json",  action="store_true", help="Output raw JSON instead of formatted diff")
    args = parser.parse_args()

    result = diff(Path(args.old), Path(args.new), args.sheet)

    if args.json:
        print(json.dumps(result, indent=2))
    else:
        print_diff(result)


if __name__ == "__main__":
    main()
